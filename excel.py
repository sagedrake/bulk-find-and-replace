import openpyxl as xl
import eventlog
from word import str_find_and_replace


def xlsx_find_and_replace(filepath, old_word, new_word):
    """
    Replace instances of the given old word with the given new word in sheet titles, cell contents of given Excel file
    :param filepath: The name of the file to be edited
    :param old_word: The word to be searched for, assumed to be capitalized (e.g. 'Shark' not 'shark')
    :param new_word: The word to replace old_word with, assumed to be capitalized (e.g. 'Whale' not 'WHALE')
    :return: None
    """
    try:
        wb = xl.load_workbook(filepath)

        for sheet in wb.worksheets:
            sheet.title = str_find_and_replace(sheet.title, old_word, new_word)
            for row in range(1, sheet.max_row + 1):
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row, col)
                    if isinstance(cell.value, str):
                        cell.value = str_find_and_replace(cell.value, old_word, new_word)

        wb.save(filepath)
        eventlog.log_event("Replaced instances of " + old_word + " within" + filepath)
    except PermissionError:
        eventlog.log_event("ERROR: Could not edit contents of " + filepath + " because this file is already in use.")



